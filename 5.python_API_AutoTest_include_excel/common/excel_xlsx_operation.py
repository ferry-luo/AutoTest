# encoding:utf-8
# @Time:2020/10/17

# @Author:Ferry

# @colleges_class:湖南理工学院_电子14-2BF
from openpyxl import *
from openpyxl.styles import Font


# Excel数据的读取、获取单元格内容、获取一行数据和数据写入等基础操作
class ExcelOperate(object):
    def __init__(self, file_path, sheet_name):
        self.file_path = file_path
        self.sheet_name = sheet_name
        self.excel = self.get_all_excel()
        self.sheet_data = self.get_sheet_data(sheet_name)

    # 获取整个excel数据
    def get_all_excel(self):
        """
        :return:
        """
        # excel = xlrd.open_workbook(self.file_path,formatting_info=True)
        excel = load_workbook(self.file_path)
        return excel

    # 获取excel中的某个表数据
    def get_sheet_data(self, sheet_name=0):
        """
        :param sheet_name: 表名，可以是下标也可以是名称
        :return:
        """
        # if isinstance(sheet_name, int):
        #     sheet = self.excel.sheet_by_index(sheet_name)
        # else:
        #     sheet = self.excel.sheet_by_name(sheet_name)
        sheet = self.excel[sheet_name]
        return sheet

    # 获取excel行数
    def get_rows(self):
        """
        :return:
        """

        # return self.sheet_data.nrows
        return self.sheet_data.max_row

    # 获取excel列数
    def get_cols(self):
        # return self.sheet_data.ncols
        return self.sheet_data.max_column

    # 获取单元格内容
    def get_cell_value(self, row, column):
        """
        :param row:  行号
        :param column: 列号
        :return:
        """
        data = self.sheet_data.cell(row, column).value
        return data

    # 获取某列头的索引
    def get_col_index(self, col_name):
        col_index = None
        # 与xlrd不同，openpyxl 读写单元格时，单元格的坐标位置起始值是（1,1），即下标最小值为1，否则报错！
        for i in range(1, self.sheet_data.max_column + 1):
            if self.sheet_data.cell(1, i).value == col_name:
                col_index = i - 1
                break
        return col_index

    # 获取一行的数据
    def get_row_values(self, row):
        """
        :param row:
        """
        row_data = []
        for i in range(1, self.sheet_data.max_column + 1):
            cell_value = self.sheet_data.cell(row=row, column=i).value
            row_data.append(cell_value)
        return row_data

    # 获取一列的数据
    def get_col_values(self, col):
        """
        :param col:
        """
        col_data = []
        for i in range(1, self.sheet_data.max_row + 1):
            cell_value = self.sheet_data.cell(row=i, column=col).value
            col_data.append(cell_value)
        return col_data

    # 数据写入
    def write_value(self, row, column, value, value_color):
        """
        :param row: 行
        :param column: 列
        :param value: 数据
        :return:
        """
        self.sheet_data.cell(row=row + 1, column=column + 1).value = value
        self.sheet_data.cell(row=row + 1, column=column + 1).font = Font(color=value_color)
        self.excel.save(self.file_path)
